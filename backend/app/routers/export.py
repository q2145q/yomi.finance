"""Роутер для загрузки шаблона бюджета в проект и экспорта в Excel."""
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.budget import BudgetLine
from app.models.user import ProjectUser
from app.core.budget_template import build_flat_lines
from app.routers.deps import CurrentUser

router = APIRouter(tags=["budget-template"])


@router.post("/projects/{project_id}/budget/from-template", status_code=201)
async def load_template(project_id: uuid.UUID, current_user: CurrentUser, db: AsyncSession = Depends(get_db)):
    """Загружает стандартный шаблон бюджета в проект (удаляет существующие статьи)."""
    # Проверяем доступ
    if not current_user.is_superadmin:
        pu_result = await db.execute(
            select(ProjectUser).where(ProjectUser.project_id == project_id, ProjectUser.user_id == current_user.id)
        )
        pu = pu_result.scalar_one_or_none()
        if not pu or pu.role not in ("PRODUCER", "LINE_PRODUCER"):
            raise HTTPException(status_code=403, detail="Только продюсер может загружать шаблон")

    # Удаляем существующие статьи
    await db.execute(delete(BudgetLine).where(BudgetLine.project_id == project_id))

    # Генерируем плоский список из шаблона
    lines_data = build_flat_lines(project_id=project_id)

    # Bulk insert
    lines = [BudgetLine(**d) for d in lines_data]
    db.add_all(lines)
    await db.commit()

    return {"message": f"Загружено {len(lines)} статей бюджета", "count": len(lines)}


@router.get("/projects/{project_id}/budget/export")
async def export_budget_excel(project_id: uuid.UUID, current_user: CurrentUser, db: AsyncSession = Depends(get_db)):
    """Экспорт бюджета в Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")

    lines_result = await db.execute(
        select(BudgetLine).where(BudgetLine.project_id == project_id).order_by(BudgetLine.sort_order)
    )
    lines = list(lines_result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Бюджет"

    # Заголовки
    headers = ["Код", "Статья", "Ед.изм.", "Кол-во ед.", "Ставка", "Кол-во", "Итого нетто", "Лимит"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)

    # Данные
    for row, line in enumerate(lines, 2):
        indent = "  " * line.level
        ws.cell(row=row, column=1, value=line.code)
        ws.cell(row=row, column=2, value=indent + line.name)
        ws.cell(row=row, column=3, value=line.unit)
        ws.cell(row=row, column=4, value=line.quantity_units)
        ws.cell(row=row, column=5, value=line.rate)
        ws.cell(row=row, column=6, value=line.quantity)
        ws.cell(row=row, column=7, value=line.rate * line.quantity)
        ws.cell(row=row, column=8, value=line.limit_amount)

        if line.level == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"budget_{project_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
